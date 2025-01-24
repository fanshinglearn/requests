# 應該存到資料庫
import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from xvideos_class import Xvideos
from pornhub_class import Pornhub
from missav_class import Missav
from hanime1_class import Hanime1
import my_function as mf


def excel_header(ws1, ws2):
    if ws1:
        # 標題
        header = ['標題', '存在', '網址', '番號', '縮圖', '簡圖', '集數', '頻道', '模特']
        ws1.append(header)
        # 凍結窗格
        ws1.freeze_panes = 'B2'
        # 左右置中 + 粗體
        for col in range(1, len(header)+1):
            char = get_column_letter(col)
            ws1[f'{char}1'].alignment = Alignment(horizontal='center')
            ws1[f'{char}1'].font = Font(bold=True)
    if ws2:
        # 標題
        ws2.append(['縮圖', '標題'])
        # 凍結窗格
        ws2.freeze_panes = 'B2'
        # 左右置中 + 粗體
        # A B
        for col in range(1, 3):
            char = get_column_letter(col)
            ws2[f'{char}1'].alignment = Alignment(horizontal='center')
            ws2[f'{char}1'].font = Font(bold=True, size=20)
        # 欄寬
        ws2.column_dimensions['B'].width = 170


def get_website(url):
    if 'xvideos' in url:
        return 'xvideos'
    elif 'pornhub' in url:
        return 'pornhub'
    elif 'missav' in url:
        return 'missav'
    elif 'hanime1' in url:
        return 'hanime1'
    elif 'https://' not in url:
        return 'num'
    else:
        raise ValueError('input.xlsx website 輸入錯誤!')


def create_website_video(website, url):
    if website == 'xvideos':
        return Xvideos(url)
    elif website == 'pornhub':
        return Pornhub(url)
    elif website == 'missav':
        return Missav(url)
    elif website == 'hanime1':
        return Hanime1(url)
    elif website == 'num':
        video = Missav(f'https://missav.com/{str(url).lower()}')
        if video.is_exist:
            video_uncensored_leak = Missav(f'https://missav.com/{str(url).lower()}-uncensored-leak')
            return video_uncensored_leak if video_uncensored_leak.is_exist else video
        else:
            return video
    else:
        raise ValueError('website 輸入錯誤!')


def excel_content(ws1, ws2, website, input_list):
    for row, url in enumerate(input_list, start=2):
        count = row - 1
        print(f'( {count} / {len(input_list)} ) . . .')
        # 不同網站
        video = create_website_video(website, url)
        if ws1:
            # 影片資訊
            if video.is_exist:
                # 內容
                content = [video.title, video.is_exist, video.url, video.num,
                            video.thumbnail_url, video.thumbnail_side_url,
                            video.episode, video.channel]
                ws1.append(content)
                # 超連結
                ws1[f'A{row}'].hyperlink = video.url
                ws1[f'E{row}'].hyperlink = video.thumbnail_url
                ws1[f'F{row}'].hyperlink = video.thumbnail_side_url
                ws1[f'H{row}'].hyperlink = video.channel_url
                # 多個 model
                if video.model:
                    for i in range(len(video.model)):
                        col = i + len(content) + 1
                        char = get_column_letter(col)
                        ws1[f'{char}{row}'] = video.model[i]
                        ws1[f'{char}{row}'].hyperlink = video.model_url[i]
            else:
                ws1.append([video.title, video.is_exist, video.url, video.num])
                ws1[f'A{row}'].hyperlink = video.url
        # 影片縮圖
        if ws2:
            if video.is_exist:
                img_filename = f'{count}.jpg'
                mf.MyFunction.download_img(video.thumbnail_url, img_filename)
                mf.OpenpyxlFunction.add_thumbnail_to_excel(ws2, img_filename, f'A{row}')
            # 標題
            ws2[f'B{row}'] = video.title
            ws2[f'B{row}'].font = Font(size=20)
            ws2[f'B{row}'].alignment = Alignment(vertical='center')
            # 超連結
            ws2[f'B{row}'].hyperlink = video.url


def videos_imformation(imformation=True, thumbnail=True):
    # 創建並切換資料夾
    mf.MyFunction.mkdir_and_chdir_folder('porn_excel')
    # 讀取 excel
    input_list = mf.OpenpyxlFunction.load_input('input.xlsx')
    website = get_website(input_list[0])
    # 創建 excel
    wb = Workbook()
    wb.remove(wb.active)
    if imformation:
        wb.create_sheet(f'{website} 詳細資訊')
        ws1 = wb[f'{website} 詳細資訊']
    else:
        ws1 = None
    if thumbnail:
        wb.create_sheet(f'{website} 縮圖')
        ws2 = wb[f'{website} 縮圖']
    else:
        ws2 = None
    # 標題
    excel_header(ws1, ws2)
    # 內容
    excel_content(ws1, ws2, website, input_list)
    # 存檔
    exc_path = 'output.xlsx'
    wb.save(exc_path)
    # 開啟 excel
    os.startfile(exc_path)
    # 刪除下載的照片
    if thumbnail:
        mf.MyFunction.delete_images_in_folder()
    print('_' * 30)
    print(f'已新增 {len(input_list)} 部影片詳細資訊')


if __name__ == '__main__':
    # videos_imformation()
    videos_imformation(imformation=False, thumbnail=True)
