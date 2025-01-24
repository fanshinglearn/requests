import os

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from pytube import Playlist
from pytube import YouTube

class MyFunction:

    @staticmethod
    def mkdir_folder(folder_path):
        # 如果沒有資料夾就創建一個
        if not os.path.isdir(folder_path):
            os.mkdir(folder_path)
            print(f'正在創建 {folder_path} 資料夾 . . .')

    @staticmethod
    def mkdir_and_chdir_folder(folder_path):
        # 如果沒有資料夾就創建一個 並切換目錄到資料夾
        if not os.path.isdir(folder_path):
            os.mkdir(folder_path)
            print(f'正在創建 {folder_path} 資料夾 . . .')
        os.chdir(folder_path)
        print(f'正在切換到 {folder_path} 資料夾 . . .')

    @staticmethod
    def download_img(img_url, file_name):
        # 下載圖片
        try:
            print(f'正在下載圖片 {file_name} . . .')
            img = requests.get(img_url)
            with open(file_name, 'wb') as file:
                    file.write(img.content)
        except Exception as e:
            print(f'下載圖片錯誤： {e}')

    @staticmethod
    def delete_images_in_folder(folder_path = '.'):
        try:
            # 確保資料夾路徑存在
            if not os.path.isdir(folder_path):
                print(f'資料夾 "{folder_path}" 不存在。')
                return
            # 獲取資料夾內的所有檔案
            file_list = os.listdir(folder_path)
            # 迭代處理檔案
            for file_name in file_list:
                file_path = os.path.join(folder_path, file_name)
                # 檢查是否是圖片
                if file_name.endswith(('.jpg', '.png')):
                    os.remove(file_path)
                    print(f'已刪除圖片： {file_name}')
            print('所有圖片已成功刪除。')
        except Exception as e:
            print('刪除圖片過程中發生錯誤： ', e)

    def _delete_images_walk_folder(folder_path = '.'):
        # 刪除資料夾內的所有照片
        # 包括子資料夾
        try:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    if file.endswith(('.png', '.jpg')):
                        file_path = os.path.join(root, file)
                        os.remove(file_path)
                        print(f'Deleted: {file_path}')
            print('All images deleted successfully.')
        except Exception as e:
            print(f'刪除照片時發生錯誤 : {e}')


class OpenpyxlFunction:

    @staticmethod
    def load_input(path):
        wb = load_workbook(path)
        ws = wb.active
        return [ws[f'A{row}'].value for row in range(1, ws.max_row + 1)]

    @staticmethod
    def add_thumbnail_to_excel(ws, img_file_name, cell):
        # 添加 Youtube影片縮圖 到 excel 儲存格
        try:
            img = Image(img_file_name)
            # 放大儲存格
            # 以 (1280 * 720) / 4 為基準，每隔 0.1 
            # x = 45.7
            ws.column_dimensions[cell[0]].width = 45.7
            # 135.0 <= x <= 135.7
            ws.row_dimensions[int(cell[1:])].height = 135.3
            # 縮小圖片
            x = img.width / 320
            # Test 前三個是 YouTube
            # width_list = [1280, 640, 480, 600, 690]
            # if img.width not in width_list:
            #     raise Exception('新影片縮圖尺寸' + '!' * 100)
            # 圖片長寬
            img.width /= x
            img.height /= x
            # 添加圖片
            ws.add_image(img, cell)
        except Exception as e:
            print(f'添加圖片 {img_file_name} 時出現錯誤： {e}')

class PytubeFunction:

    @staticmethod
    def playlist_url_to_video_objects(playlist_url):
        playlist = Playlist(playlist_url)
        videos = [YouTube(url) for url in playlist.video_urls]
        return videos
    
    @staticmethod
    def playlist_to_video_objects(playlist):
        videos = [YouTube(url) for url in playlist.video_urls]
        return videos

    @staticmethod
    def common_video_urls(playlist_url1, playlist_url2):
        # set 執行速度較快
        playlist1 = Playlist(playlist_url1)
        playlist2 = Playlist(playlist_url2)

        videos1 = set(playlist1.video_urls)
        videos2 = set(playlist2.video_urls)

        common_urls = videos1 & videos2
        return common_urls

    @classmethod
    def common_video_objects(cls, playlist_url1, playlist_url2):
        common_urls = cls.common_video_urls(playlist_url1, playlist_url2)
        common_videos = [YouTube(url) for url in common_urls]
        return common_videos
